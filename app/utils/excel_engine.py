"""
Excel Engine - Import/Export Excel files with diff comparison.
"""

import os
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# Excel column mapping
EXCEL_COLUMNS = {
    'A': 'code',           # Số hiệu
    'B': 'category_name',  # Loại hạng mục
    'C': 'pier',           # Trụ
    'D': 'span',           # Nhịp
    'E': 'segment',        # Đốt/Đợt
    'F': 'volume',         # Khối lượng
    'G': 'unit',           # Đơn vị
    'H': 'unit_price',     # Đơn giá
    'I': 'total_value',    # Tổng giá trị
    'J': 'status',         # Trạng thái
    'K': 'completed_at',   # Ngày hoàn thành
    'L': 'notes',          # Ghi chú
}

STATUS_TEXT = {
    0: "Chưa",
    1: "Đang",
    2: "Xong"
}

STATUS_FROM_TEXT = {
    "chưa": 0, "chua": 0,
    "đang": 1, "dang": 1,
    "xong": 2, "hoàn thành": 2, "hoan thanh": 2, "done": 2
}


def parse_status(value) -> int:
    """Parse status from various formats."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value) if 0 <= value <= 2 else 0
    
    text = str(value).lower().strip()
    return STATUS_FROM_TEXT.get(text, 0)


def read_excel(filepath: str) -> List[Dict]:
    """
    Read blocks from Excel file.
    Returns list of block dictionaries.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    
    blocks = []
    
    # Skip header row
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row[0]:
            continue
        
        block = {
            'code': str(row[0]).strip() if row[0] else '',
            'category_name': str(row[1]).strip() if row[1] else '',
            'pier': str(row[2]).strip() if row[2] else None,
            'span': str(row[3]).strip() if row[3] else None,
            'segment': str(row[4]).strip() if row[4] else None,
            'volume': float(row[5]) if row[5] is not None else None,
            'unit': str(row[6]).strip() if row[6] else 'm³',
            'unit_price': float(row[7]) if row[7] is not None else None,
            'total_value': float(row[8]) if row[8] is not None else None,
            'status': parse_status(row[9]),
            'completed_at': str(row[10]) if row[10] else None,
            'notes': str(row[11]).strip() if len(row) > 11 and row[11] else None,
        }
        
        # Calculate total if not provided
        if block['total_value'] is None and block['volume'] and block['unit_price']:
            block['total_value'] = block['volume'] * block['unit_price']
        
        blocks.append(block)
    
    wb.close()
    return blocks


def export_excel(blocks: List[Dict], filepath: str, project_name: str = ""):
    """
    Export blocks to Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Số hiệu", "Loại hạng mục", "Trụ", "Nhịp", "Đốt/Đợt",
        "Khối lượng", "Đơn vị", "Đơn giá", "Tổng giá trị",
        "Trạng thái", "Ngày HT", "Ghi chú"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row_idx, block in enumerate(blocks, 2):
        ws.cell(row=row_idx, column=1, value=block.get('code', ''))
        ws.cell(row=row_idx, column=2, value=block.get('category_name', ''))
        ws.cell(row=row_idx, column=3, value=block.get('pier', ''))
        ws.cell(row=row_idx, column=4, value=block.get('span', ''))
        ws.cell(row=row_idx, column=5, value=block.get('segment', ''))
        ws.cell(row=row_idx, column=6, value=block.get('volume'))
        ws.cell(row=row_idx, column=7, value=block.get('unit', 'm³'))
        ws.cell(row=row_idx, column=8, value=block.get('unit_price'))
        ws.cell(row=row_idx, column=9, value=block.get('total_value'))
        ws.cell(row=row_idx, column=10, value=STATUS_TEXT.get(block.get('status', 0), 'Chưa'))
        ws.cell(row=row_idx, column=11, value=block.get('completed_at', ''))
        ws.cell(row=row_idx, column=12, value=block.get('notes', ''))
        
        # Apply border
        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).border = border
    
    # Column widths
    col_widths = [12, 18, 8, 8, 10, 12, 8, 15, 18, 10, 12, 25]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Save
    wb.save(filepath)
    wb.close()


def compare_blocks(
    excel_blocks: List[Dict],
    db_blocks: List[Dict]
) -> Dict[str, List]:
    """
    Compare Excel blocks with database blocks.
    
    Returns:
        {
            'new': [blocks in Excel but not in DB],
            'changed': [(excel_block, db_block, changed_fields)],
            'deleted': [blocks in DB but not in Excel]
        }
    """
    # Index DB blocks by code
    db_by_code = {b['code']: b for b in db_blocks}
    excel_codes = {b['code'] for b in excel_blocks}
    db_codes = set(db_by_code.keys())
    
    result = {
        'new': [],
        'changed': [],
        'deleted': []
    }
    
    # Find new blocks
    for block in excel_blocks:
        if block['code'] not in db_codes:
            result['new'].append(block)
        else:
            # Check for changes
            db_block = db_by_code[block['code']]
            changed_fields = []
            
            compare_fields = [
                'category_name', 'pier', 'span', 'segment',
                'volume', 'unit', 'unit_price', 'status', 'notes'
            ]
            
            for field in compare_fields:
                excel_val = block.get(field)
                db_val = db_block.get(field)
                
                # Normalize None and empty string
                if excel_val == '':
                    excel_val = None
                if db_val == '':
                    db_val = None
                
                # Compare with tolerance for floats
                if isinstance(excel_val, float) and isinstance(db_val, float):
                    if abs(excel_val - db_val) > 0.001:
                        changed_fields.append(field)
                elif excel_val != db_val:
                    changed_fields.append(field)
            
            if changed_fields:
                result['changed'].append((block, db_block, changed_fields))
    
    # Find deleted blocks
    for code in db_codes:
        if code not in excel_codes:
            result['deleted'].append(db_by_code[code])
    
    return result


def create_diff_report(
    diff_result: Dict[str, List],
    filepath: str
):
    """
    Create an Excel diff report showing changes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "So sánh thay đổi"
    
    # Styles
    new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
    changed_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
    deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
    header_font = Font(bold=True)
    
    row = 1
    
    # New blocks section
    if diff_result['new']:
        ws.cell(row=row, column=1, value=f"🆕 BLOCKS MỚI ({len(diff_result['new'])})").font = header_font
        row += 1
        
        # Headers
        for col, header in enumerate(['Số hiệu', 'Loại', 'Trụ', 'Khối lượng', 'Đơn giá'], 1):
            ws.cell(row=row, column=col, value=header).font = header_font
        row += 1
        
        for block in diff_result['new']:
            cells = [
                block.get('code', ''),
                block.get('category_name', ''),
                block.get('pier', ''),
                block.get('volume'),
                block.get('unit_price')
            ]
            for col, val in enumerate(cells, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = new_fill
            row += 1
        
        row += 1
    
    # Changed blocks section
    if diff_result['changed']:
        ws.cell(row=row, column=1, value=f"✏️ BLOCKS THAY ĐỔI ({len(diff_result['changed'])})").font = header_font
        row += 1
        
        for excel_block, db_block, changed_fields in diff_result['changed']:
            ws.cell(row=row, column=1, value=excel_block.get('code', '')).font = header_font
            ws.cell(row=row, column=2, value=f"Thay đổi: {', '.join(changed_fields)}")
            for col in range(1, 3):
                ws.cell(row=row, column=col).fill = changed_fill
            row += 1
        
        row += 1
    
    # Deleted blocks section
    if diff_result['deleted']:
        ws.cell(row=row, column=1, value=f"🗑️ BLOCKS BỊ XÓA ({len(diff_result['deleted'])})").font = header_font
        row += 1
        
        for block in diff_result['deleted']:
            cell = ws.cell(row=row, column=1, value=block.get('code', ''))
            cell.fill = deleted_fill
            row += 1
    
    # Adjust column width
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    
    wb.save(filepath)
    wb.close()


def create_sample_excel(filepath: str, project_type: str = "bridge"):
    """
    Create a sample Excel template for bridge/road projects.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    
    # Headers
    headers = [
        "Số hiệu", "Loại hạng mục", "Trụ", "Nhịp", "Đốt/Đợt",
        "Khối lượng", "Đơn vị", "Đơn giá", "Tổng giá trị",
        "Trạng thái", "Ngày HT", "Ghi chú"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Sample data for bridge
    if project_type == "bridge":
        sample_data = [
            # Cọc khoan nhồi - Trụ T1
            ["C1-T1", "Cọc khoan nhồi", "T1", "", "", 45.5, "m³", 3500000, 159250000, "Xong", "2024-01-15", "Cọc D1200"],
            ["C2-T1", "Cọc khoan nhồi", "T1", "", "", 45.5, "m³", 3500000, 159250000, "Xong", "2024-01-16", ""],
            ["C3-T1", "Cọc khoan nhồi", "T1", "", "", 45.5, "m³", 3500000, 159250000, "Đang", "", ""],
            ["C4-T1", "Cọc khoan nhồi", "T1", "", "", 45.5, "m³", 3500000, 159250000, "Chưa", "", ""],
            
            # Cọc khoan nhồi - Trụ T2
            ["C1-T2", "Cọc khoan nhồi", "T2", "", "", 45.5, "m³", 3500000, 159250000, "Chưa", "", ""],
            ["C2-T2", "Cọc khoan nhồi", "T2", "", "", 45.5, "m³", 3500000, 159250000, "Chưa", "", ""],
            
            # Bệ trụ
            ["BT1-D1", "Bệ trụ", "T1", "", "Đợt 1", 120, "m³", 2800000, 336000000, "Chưa", "", ""],
            ["BT1-D2", "Bệ trụ", "T1", "", "Đợt 2", 80, "m³", 2800000, 224000000, "Chưa", "", ""],
            
            # Thân trụ
            ["TT1-D1", "Thân trụ", "T1", "", "Đốt 1", 35, "m³", 3200000, 112000000, "Chưa", "", ""],
            ["TT1-D2", "Thân trụ", "T1", "", "Đốt 2", 35, "m³", 3200000, 112000000, "Chưa", "", ""],
            ["TT1-D3", "Thân trụ", "T1", "", "Đốt 3", 35, "m³", 3200000, 112000000, "Chưa", "", ""],
            
            # Xà mũ
            ["XM1", "Xà mũ", "T1", "", "", 25, "m³", 3000000, 75000000, "Chưa", "", ""],
            
            # Dầm
            ["D-N1", "Dầm", "", "N1", "", 180, "m³", 4500000, 810000000, "Chưa", "", "Dầm Super-T"],
        ]
    else:
        # Road project sample
        sample_data = [
            ["KC1", "Kết cấu mặt đường", "", "", "Km0+000-Km0+500", 2500, "m²", 450000, 1125000000, "Chưa", "", ""],
            ["KC2", "Kết cấu mặt đường", "", "", "Km0+500-Km1+000", 2500, "m²", 450000, 1125000000, "Chưa", "", ""],
        ]
    
    for row_idx, data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Column widths
    col_widths = [12, 18, 8, 8, 15, 12, 8, 15, 18, 10, 12, 25]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    wb.save(filepath)
    wb.close()
