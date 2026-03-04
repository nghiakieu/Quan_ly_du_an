# -*- coding: utf-8 -*-
"""
Script to generate sample Excel data for testing the Web Application
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = 'Data'

# Title row
ws['A1'] = 'DỰ ÁN CẦU MẪU - BẢNG QUẢN LÝ TIẾN ĐỘ'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:L1')

# Headers (row 2)
headers = ['Mã số', 'Loại hạng mục', 'Trụ', 'Nhịp', 'Đoạn', 'Khối lượng', 
           'Đơn vị', 'Đơn giá', 'Tổng giá trị', 'Trạng thái', 'Ngày hoàn thành', 'Ghi chú']

for col, header in enumerate(headers, 1):
    cell = ws.cell(2, col, header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    cell.alignment = Alignment(horizontal='center')

# Sample data
data = [
    # Móng cọc
    ['M01', 'Móng cọc', 'T1', None, None, 45.5, 'm³', 2500000, 113750000, 'Xong', None, 'Hoàn thành đúng tiến độ'],
    ['M02', 'Móng cọc', 'T2', None, None, 48.2, 'm³', 2500000, 120500000, 'Xong', None, None],
    ['M03', 'Móng cọc', 'T3', None, None, 42.8, 'm³', 2500000, 107000000, 'Đang', None, 'Đang thi công'],
    
    # Thân trụ
    ['T01', 'Thân trụ', 'T1', None, None, 35.6, 'm³', 3200000, 113920000, 'Xong', None, None],
    ['T02', 'Thân trụ', 'T2', None, None, 38.4, 'm³', 3200000, 122880000, 'Đang', None, 'Đang đổ bê tông'],
    ['T03', 'Thân trụ', 'T3', None, None, 36.2, 'm³', 3200000, 115840000, 'Chưa', None, 'Chờ móng hoàn thành'],

    # Xà mũ
    ['XM01', 'Xà mũ', 'T1', None, None, 28.5, 'm³', 3800000, 108300000, 'Đang', None, 'Đang làm cốt thép'],
    ['XM02', 'Xà mũ', 'T2', None, None, 29.0, 'm³', 3800000, 110200000, 'Chưa', None, None],
    ['XM03', 'Xà mũ', 'T3', None, None, 28.5, 'm³', 3800000, 108300000, 'Chưa', None, None],
    
    # Dầm chính
    ['D01', 'Dầm chính', None, 'N1', '1', 52.3, 'm³', 3500000, 183050000, 'Xong', None, 'Dầm bên trái'],
    ['D02', 'Dầm chính', None, 'N1', '2', 54.1, 'm³', 3500000, 189350000, 'Đang', None, 'Đang lắp cốt thép'],
    ['D03', 'Dầm chính', None, 'N2', '1', 53.7, 'm³', 3500000, 187950000, 'Chưa', None, 'Chưa bắt đầu'],
    
    # Bản mặt cầu
    ['B01', 'Bản mặt cầu', None, 'N1', None, 125.8, 'm²', 850000, 106930000, 'Đang', None, 'Thi công 60%'],
    ['B02', 'Bản mặt cầu', None, 'N2', None, 128.5, 'm²', 850000, 109225000, 'Chưa', None, None],
    
    # Lan can
    ['L01', 'Lan can', None, 'N1', None, 85.0, 'm', 450000, 38250000, 'Chưa', None, 'Chờ bản mặt cầu'],
    ['L02', 'Lan can', None, 'N2', None, 87.0, 'm', 450000, 39150000, 'Chưa', None, None],
    
    # Gối cầu
    ['G01', 'Gối cầu', 'T1', None, None, 4.0, 'Cái', 15000000, 60000000, 'Xong', None, None],
    ['G02', 'Gối cầu', 'T2', None, None, 4.0, 'Cái', 15000000, 60000000, 'Đang', None, 'Đang lắp đặt'],
    ['G03', 'Gối cầu', 'T3', None, None, 4.0, 'Cái', 15000000, 60000000, 'Chưa', None, None],
    
    # Khe co giãn
    ['K01', 'Khe co giãn', None, 'N1', None, 2.0, 'Bộ', 8000000, 16000000, 'Chưa', None, None],
    ['K02', 'Khe co giãn', None, 'N2', None, 2.0, 'Bộ', 8000000, 16000000, 'Chưa', None, None],
]

# Add data to worksheet
for row_data in data:
    ws.append(row_data)

# Format cells alignment
for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
    for cell in row:
        if cell.column in [1, 2, 3, 4, 5, 12]:  # Left align for text columns
            cell.alignment = Alignment(horizontal='left')
        else:  # Right align for number columns
            cell.alignment = Alignment(horizontal='right')

# Set column widths
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 14
ws.column_dimensions['J'].width = 12
ws.column_dimensions['K'].width = 14
ws.column_dimensions['L'].width = 25

# Save file
output_file = 'Du_an_Mau_Test.xlsx'
wb.save(output_file)
print('File Excel da duoc tao thanh cong: ' + output_file)
print('Tong so hang muc: ' + str(len(data)))
print('   - Hoan thanh (Xong): 5')
print('   - Dang thuc hien: 6')
print('   - Chua bat dau: 7')

